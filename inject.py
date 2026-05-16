"""
Inyector unificado: toma los horarios parseados (SR + PDF) y los escribe
en las hojas individuales del Excel de Lili.

- Cada persona tiene una fuente única (SR o PDF) definida en mapping.json.
- Mantiene fórmulas/formato existente. Detecta layout dinámico de cada hoja.
- NUNCA toca el archivo original — escribe siempre en outputs/.

Uso:
    python inject.py
"""
from __future__ import annotations
import json
from pathlib import Path
from datetime import datetime, time, timedelta
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

DEFAULT_INPUT_NOMINA = Path("inputs/NOMINA I MAYO 2026.xlsx")
DEFAULT_INPUT_SR = Path("outputs/horarios_parseados.xlsx")
DEFAULT_INPUT_PDF = Path("outputs/horarios_pdf.xlsx")
DEFAULT_OUTPUT = Path("outputs/NOMINA I MAYO 2026 - PRELLENADO.xlsx")
MAPPING_FILE = Path("mapping.json")

START_ROW = 5
PERIODO_DAYS = 15

FILL_REVISION = PatternFill("solid", fgColor="FFE699")
FILL_INCOMPLETO = PatternFill("solid", fgColor="F4B084")


def load_mapping(mapping_path: Path = MAPPING_FILE) -> list[dict]:
    raw = json.loads(mapping_path.read_text(encoding="utf-8"))
    return raw["mapping"]


def hhmm_to_time(s):
    """Convierte 'HH:MM' a time. Si está vacío devuelve None (celda en blanco)."""
    if s is None or pd.isna(s):
        return None
    s = str(s).strip()
    if not s or s in ("nan", "NaT") or ":" not in s:
        return None
    h, m = s.split(":")[:2]
    return time(int(h), int(m))


def is_empty_day(row) -> bool:
    """True si TODAS las celdas del día vienen vacías (libre o ausencia)."""
    return all(
        (v is None or pd.isna(v) or str(v).strip() in ("", "nan", "NaT"))
        for v in (row["ENT_AM"], row["SAL_AM"], row["ENT_PM"], row["SAL_PM"])
    )


def detect_layout(ws) -> dict:
    """
    Detecta automáticamente layout de la hoja:
      - day_col: columna del día (A o B)
      - start_row: primera fila con día 29 (puede ser 5 o 6 según la hoja)
      - am_in, am_out, pm_in, pm_out: columnas de los 4 horarios
    """
    # 1) Buscar día 29 (primer día de la quincena) escaneando filas 4-8 y cols 1-2
    day_col = None
    start_row = None
    for r in range(4, 9):
        for col in (1, 2):
            v = ws.cell(row=r, column=col).value
            if isinstance(v, (int, float)) and int(v) == 29:
                day_col, start_row = col, r
                break
        if start_row:
            break
    if not start_row:
        raise ValueError(f"No encontré el día 29 en la hoja '{ws.title}'")

    # 2) Detectar columnas con time. Escanear desde start_row varias filas.
    pure_time_cols: list[int] = []
    for col in range(day_col + 1, 15):
        for r in range(start_row, start_row + 6):
            v = ws.cell(row=r, column=col).value
            if isinstance(v, time):
                pure_time_cols.append(col)
                break
            if isinstance(v, str) and v.startswith("="):
                break

    if len(pure_time_cols) == 3:
        am_in, am_out, pm_in = pure_time_cols
        pm_out = pm_in + 1
    elif len(pure_time_cols) >= 4:
        am_in, am_out, pm_in, pm_out = pure_time_cols[:4]
    else:
        raise ValueError(
            f"Layout de '{ws.title}': encontré {len(pure_time_cols)} cols de tiempo ({pure_time_cols})"
        )
    return {
        "day_col": day_col, "start_row": start_row,
        "am_in": am_in, "am_out": am_out, "pm_in": pm_in, "pm_out": pm_out,
    }


def load_source_data(sr_path: Path, pdf_path: Path) -> dict[str, pd.DataFrame]:
    """Carga los dos archivos de horarios parseados y los devuelve en dict."""
    sources = {}
    if sr_path.exists():
        df = pd.read_excel(sr_path)
        df["Fecha_dt"] = pd.to_datetime(df["Fecha"])
        sources["SR"] = df
    if pdf_path.exists():
        df = pd.read_excel(pdf_path)
        df["Fecha_dt"] = pd.to_datetime(df["Fecha"])
        sources["PDF"] = df
    return sources


def inject(
    input_nomina: Path = DEFAULT_INPUT_NOMINA,
    input_sr: Path = DEFAULT_INPUT_SR,
    input_pdf: Path = DEFAULT_INPUT_PDF,
    output: Path = DEFAULT_OUTPUT,
    log_fn=print,
) -> dict:
    """Inyecta horarios parseados en el Excel de Lili. Devuelve dict con resumen."""
    if not input_nomina.exists():
        raise FileNotFoundError(f"Falta {input_nomina}")

    mapping = load_mapping()
    sources = load_source_data(input_sr, input_pdf)
    if not sources:
        raise RuntimeError("Sin fuentes de datos. Corre parser.py y parser_pdf.py primero.")

    # Determinar período (mínimo de cualquier fuente disponible)
    all_dates = pd.concat([df["Fecha_dt"] for df in sources.values()])
    start = all_dates.min()
    valid_dates = [start + timedelta(days=d) for d in range(PERIODO_DAYS)]

    log_fn(f"Cargando {input_nomina.name}...")
    wb = load_workbook(input_nomina)
    log_fn(f"Hojas disponibles: {len(wb.sheetnames)} | Periodo: {start.date()} (+{PERIODO_DAYS} dias)")

    log = []
    detalle_por_persona = []
    total_dias = 0
    total_rev = 0
    for entry in mapping:
        sheet_name = entry["sheet"]
        clock_name = entry["clock_name"]
        source = entry["source"]
        if sheet_name not in wb.sheetnames:
            log.append(f"  [X] Hoja '{sheet_name}' no existe en el workbook")
            continue
        if source not in sources:
            log.append(f"  [X] Fuente '{source}' no disponible para {sheet_name}")
            continue

        df = sources[source]
        person = df[df["Nombre"].astype(str).str.strip() == clock_name.strip()]
        person = person[person["Fecha_dt"].isin(valid_dates)]
        if person.empty:
            log.append(f"  [!] {clock_name} ({source}) no encontrado en el reporte")
            continue

        ws = wb[sheet_name]
        try:
            layout = detect_layout(ws)
        except ValueError as e:
            log.append(f"  [X] {sheet_name}: {e}")
            continue

        person = person.sort_values("Fecha_dt").reset_index(drop=True)
        dias_escritos = 0
        revision_dias = 0
        for idx, row in person.iterrows():
            xl_row = layout["start_row"] + idx
            if is_empty_day(row):
                # Día sin fichajes (libre o ausencia): 00:00 en las 4 celdas, como hace Lili.
                for col_key in ("am_in", "am_out", "pm_in", "pm_out"):
                    ws.cell(row=xl_row, column=layout[col_key]).value = time(0, 0)
            else:
                # Día con fichajes: las celdas vacías se quedan VACÍAS (None) para no
                # romper las fórmulas con cálculos negativos.
                ws.cell(row=xl_row, column=layout["am_in"]).value = hhmm_to_time(row["ENT_AM"])
                ws.cell(row=xl_row, column=layout["am_out"]).value = hhmm_to_time(row["SAL_AM"])
                ws.cell(row=xl_row, column=layout["pm_in"]).value = hhmm_to_time(row["ENT_PM"])
                ws.cell(row=xl_row, column=layout["pm_out"]).value = hhmm_to_time(row["SAL_PM"])

            estado = str(row.get("Estado", ""))
            fill = None
            if estado == "REVISION":
                fill = FILL_REVISION
                revision_dias += 1
            elif estado == "INCOMPLETO":
                fill = FILL_INCOMPLETO
                revision_dias += 1
            if fill:
                for col_key in ("am_in", "am_out", "pm_in", "pm_out"):
                    ws.cell(row=xl_row, column=layout[col_key]).fill = fill
            dias_escritos += 1

        total_dias += dias_escritos
        total_rev += revision_dias
        marker = "[REV]" if revision_dias else "[OK] "
        log.append(f"  {marker} {clock_name:15s} ({source:3s}) -> {sheet_name:12s} | {dias_escritos} dias | {revision_dias} revision")
        detalle_por_persona.append({
            "Colaborador": clock_name,
            "Hoja": sheet_name.strip(),
            "Reloj": source,
            "Días escritos": dias_escritos,
            "Para revisar": revision_dias,
        })

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)

    for line in log:
        log_fn(line)
    log_fn(f"TOTAL: {total_dias} dias-persona inyectados | {total_rev} marcados para revision")

    return {
        "total_dias": total_dias,
        "total_revision": total_rev,
        "detalle": detalle_por_persona,
        "output_path": str(output),
    }
if __name__ == "__main__":
    result = inject()
    print(f"\nOutput: {result['output_path']}")
    print("El archivo original NO fue modificado.")
