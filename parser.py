"""
Parser de fichajes del reloj Teralog/Hikvision.
Lee el "Reporte de Asistencia" (StandardReport.xls) y transforma los strings
crudos de fichajes en pares Entrada-Salida AM/PM listos para la nómina.

Uso:
    python parser.py inputs/17_StandardReport.xls outputs/

Output:
    horarios_parseados.xlsx  -> Tabla limpia: ID, Nombre, Día, ENT-AM, SAL-AM, ENT-PM, SAL-PM
    revision_manual.xlsx     -> Solo los días que el algoritmo no resolvió con certeza
"""
from __future__ import annotations
import sys, re
from pathlib import Path
from datetime import datetime, timedelta
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# Configuracion
DEDUPE_MIN = 2          # fichajes con diferencia < 2 min se consideran duplicados
ALMUERZO_MAX_MIN = 90   # gap < 90 min entre fichajes = transicion AM->PM (descanso comida)
DAY_BREAK_MIN = 240     # gap > 4h entre dos unicos fichajes = jornada continua AM-PM
NULL_TIMES = {"00:00"}  # fichajes a ignorar


def hhmm_to_min(s: str) -> int:
    h, m = s.split(":")
    return int(h) * 60 + int(m)


def parse_fichajes(raw: str | float) -> list[str]:
    """Convierte string concatenado '11:5915:0516:0422:25' en ['11:59','15:05',...]"""
    if raw is None or (isinstance(raw, float) and pd.isna(raw)):
        return []
    s = str(raw).strip()
    if not s:
        return []
    # Cada timestamp ocupa 5 caracteres HH:MM
    matches = re.findall(r"\d{2}:\d{2}", s)
    # Filtrar nulos y deduplicar muy cercanos
    cleaned: list[str] = []
    for t in matches:
        if t in NULL_TIMES:
            continue
        if cleaned and abs(hhmm_to_min(t) - hhmm_to_min(cleaned[-1])) < DEDUPE_MIN:
            continue
        cleaned.append(t)
    cleaned.sort(key=hhmm_to_min)
    # Segunda pasada de dedup tras ordenar
    final: list[str] = []
    for t in cleaned:
        if final and abs(hhmm_to_min(t) - hhmm_to_min(final[-1])) < DEDUPE_MIN:
            continue
        final.append(t)
    return final


def classify_day(fichajes: list[str]) -> dict:
    """
    Asigna fichajes a slots ENT_AM, SAL_AM, ENT_PM, SAL_PM.
    Devuelve dict con slots + estado (OK / REVISION / LIBRE / INCOMPLETO).
    """
    out = {"ENT_AM": "", "SAL_AM": "", "ENT_PM": "", "SAL_PM": "",
           "n_fichajes": len(fichajes), "estado": "", "obs": ""}

    n = len(fichajes)
    if n == 0:
        out["estado"] = "LIBRE_O_AUSENCIA"
        out["obs"] = "Sin marcas. ¿Día libre o ausencia?"
        return out

    if n == 1:
        out["ENT_AM"] = fichajes[0]
        out["estado"] = "INCOMPLETO"
        out["obs"] = "Solo 1 fichaje (falta resto del día)"
        return out

    if n == 2:
        gap = hhmm_to_min(fichajes[1]) - hhmm_to_min(fichajes[0])
        if gap >= DAY_BREAK_MIN:
            # Jornada continua: entrada AM, salida PM
            out["ENT_AM"] = fichajes[0]
            out["SAL_PM"] = fichajes[1]
            out["estado"] = "OK"
            out["obs"] = f"Jornada continua (gap {gap} min)"
        else:
            # Solo un bloque corto: AM o PM segun hora
            primer = hhmm_to_min(fichajes[0])
            if primer < 13 * 60:
                out["ENT_AM"] = fichajes[0]
                out["SAL_AM"] = fichajes[1]
            else:
                out["ENT_PM"] = fichajes[0]
                out["SAL_PM"] = fichajes[1]
            out["estado"] = "REVISION"
            out["obs"] = f"Solo 2 fichajes en bloque corto (gap {gap} min)"
        return out

    if n == 3:
        # Heuristica: ver si el gap entre el 2do y 3er fichaje parece un almuerzo
        gap_23 = hhmm_to_min(fichajes[2]) - hhmm_to_min(fichajes[1])
        gap_12 = hhmm_to_min(fichajes[1]) - hhmm_to_min(fichajes[0])
        if gap_23 <= ALMUERZO_MAX_MIN and gap_23 < gap_12:
            # 2do es salida AM, 3ro es entrada PM, falta salida PM
            out["ENT_AM"] = fichajes[0]
            out["SAL_AM"] = fichajes[1]
            out["ENT_PM"] = fichajes[2]
            out["estado"] = "REVISION"
            out["obs"] = "Falta salida PM (3 fichajes)"
        elif gap_12 <= ALMUERZO_MAX_MIN and gap_12 < gap_23:
            # 1ro es salida del bloque previo? raro. Asumir: 1ro AM, 2do entrada PM, 3ro salida PM
            out["ENT_AM"] = fichajes[0]
            out["ENT_PM"] = fichajes[1]
            out["SAL_PM"] = fichajes[2]
            out["estado"] = "REVISION"
            out["obs"] = "Falta salida AM (3 fichajes)"
        else:
            # Caso ambiguo: poner el primero como ENT_AM y el último como SAL_PM
            out["ENT_AM"] = fichajes[0]
            out["SAL_PM"] = fichajes[-1]
            out["estado"] = "REVISION"
            out["obs"] = "3 fichajes, asignación ambigua"
        return out

    if n == 4:
        out["ENT_AM"] = fichajes[0]
        out["SAL_AM"] = fichajes[1]
        out["ENT_PM"] = fichajes[2]
        out["SAL_PM"] = fichajes[3]
        out["estado"] = "OK"
        return out

    # n >= 5 : buscar 4 puntos por clusters de tiempo
    # Tomar el primero y ultimo, y los dos mas cercanos al "punto medio" (transicion mediodia)
    # Estrategia: identificar el gap mas grande -> es el descanso. Antes = AM, despues = PM.
    gaps = [hhmm_to_min(fichajes[i + 1]) - hhmm_to_min(fichajes[i]) for i in range(n - 1)]
    max_gap_idx = gaps.index(max(gaps))
    am = fichajes[: max_gap_idx + 1]
    pm = fichajes[max_gap_idx + 1:]
    out["ENT_AM"] = am[0]
    out["SAL_AM"] = am[-1] if len(am) > 1 else ""
    out["ENT_PM"] = pm[0] if pm else ""
    out["SAL_PM"] = pm[-1] if len(pm) > 1 else ""
    out["estado"] = "REVISION" if n > 5 else "OK"
    out["obs"] = f"{n} fichajes -> cluster por gap mayor ({max(gaps)} min)"
    return out


def find_persons(df: pd.DataFrame) -> list[dict]:
    """
    Busca todas las filas que contienen 'ID:' en col 0 y extrae id+nombre.
    La fila siguiente contiene los fichajes diarios.
    """
    persons = []
    for i in range(len(df)):
        cell = str(df.iloc[i, 0]).strip()
        if cell == "ID:":
            id_val = df.iloc[i, 2]
            name = df.iloc[i, 10]
            persons.append({
                "row_id": i,
                "row_fichajes": i + 1,
                "id": str(int(id_val)) if pd.notna(id_val) and str(id_val).replace(".0", "").isdigit() else str(id_val),
                "nombre": str(name).strip() if pd.notna(name) else "",
            })
    return persons


def get_period(df: pd.DataFrame) -> tuple[datetime, datetime]:
    """Extrae fechas inicio/fin del header 'Periodo: 2026-04-29 ~ 2026-05-14'"""
    for i in range(min(10, len(df))):
        for j in range(min(15, df.shape[1])):
            val = str(df.iloc[i, j])
            m = re.search(r"(\d{4}-\d{2}-\d{2})\s*~\s*(\d{4}-\d{2}-\d{2})", val)
            if m:
                return datetime.strptime(m.group(1), "%Y-%m-%d"), datetime.strptime(m.group(2), "%Y-%m-%d")
    raise ValueError("No se encontró el periodo en el reporte")


def process(input_xls: Path, output_dir: Path) -> None:
    df = pd.read_excel(input_xls, sheet_name="Reporte de Asistencia", header=None)
    start, end = get_period(df)
    num_days = (end - start).days + 1
    persons = find_persons(df)
    print(f"Período: {start.date()} a {end.date()} ({num_days} días)")
    print(f"Personas detectadas: {len(persons)}")

    rows = []
    revision_rows = []
    for p in persons:
        for d in range(num_days):
            fecha = start + timedelta(days=d)
            col = d  # día 0 = col 0, día 1 = col 1...
            raw = df.iloc[p["row_fichajes"], col] if col < df.shape[1] else None
            fichajes = parse_fichajes(raw)
            res = classify_day(fichajes)
            row = {
                "ID": p["id"],
                "Nombre": p["nombre"],
                "Fecha": fecha.strftime("%Y-%m-%d"),
                "Día": fecha.strftime("%a %d"),
                "Raw": str(raw) if raw is not None and not pd.isna(raw) else "",
                "Fichajes": " | ".join(fichajes),
                "N": res["n_fichajes"],
                "ENT_AM": res["ENT_AM"],
                "SAL_AM": res["SAL_AM"],
                "ENT_PM": res["ENT_PM"],
                "SAL_PM": res["SAL_PM"],
                "Estado": res["estado"],
                "Observación": res["obs"],
            }
            rows.append(row)
            if res["estado"] in ("REVISION", "INCOMPLETO"):
                revision_rows.append(row)

    out_full = output_dir / "horarios_parseados.xlsx"
    out_rev = output_dir / "revision_manual.xlsx"
    write_xlsx(rows, out_full, highlight=True)
    write_xlsx(revision_rows, out_rev, highlight=True)
    print(f"\nGenerado: {out_full}")
    print(f"Generado: {out_rev}")
    print(f"\nResumen:")
    estados = pd.Series([r["Estado"] for r in rows]).value_counts()
    for k, v in estados.items():
        print(f"  {k}: {v}")


def write_xlsx(rows: list[dict], path: Path, highlight: bool = True) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Horarios"
    if not rows:
        wb.save(path)
        return
    headers = list(rows[0].keys())
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        ws.cell(row=1, column=c).font = Font(bold=True)
        ws.cell(row=1, column=c).fill = PatternFill("solid", fgColor="D9E1F2")
    fill_rev = PatternFill("solid", fgColor="FFE699")
    fill_inc = PatternFill("solid", fgColor="F4B084")
    fill_libre = PatternFill("solid", fgColor="E2EFDA")
    for r in rows:
        ws.append([r.get(h, "") for h in headers])
        if not highlight:
            continue
        last = ws.max_row
        estado = r.get("Estado", "")
        if estado == "REVISION":
            for c in range(1, len(headers) + 1):
                ws.cell(row=last, column=c).fill = fill_rev
        elif estado == "INCOMPLETO":
            for c in range(1, len(headers) + 1):
                ws.cell(row=last, column=c).fill = fill_inc
        elif estado == "LIBRE_O_AUSENCIA":
            for c in range(1, len(headers) + 1):
                ws.cell(row=last, column=c).fill = fill_libre
    # Auto width
    for col_idx, h in enumerate(headers, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max(10, min(40, len(h) + 2))
    wb.save(path)


if __name__ == "__main__":
    if len(sys.argv) >= 3:
        input_xls = Path(sys.argv[1])
        output_dir = Path(sys.argv[2])
    else:
        input_xls = Path("inputs/17_StandardReport.xls")
        output_dir = Path("outputs")
    output_dir.mkdir(parents=True, exist_ok=True)
    process(input_xls, output_dir)
